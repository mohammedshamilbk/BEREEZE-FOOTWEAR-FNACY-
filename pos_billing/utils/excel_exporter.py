# ============================================================
# pos_billing/utils/excel_exporter.py
# ============================================================
"""
Excel Exporter Module for Bereeze Footwear POS & Billing System.
Generates beautifully formatted Excel (.xlsx) reports with openpyxl.
Includes daily sales summaries, day closing reports, bills log, expenses, and inventory.
"""

from __future__ import annotations

import csv
import logging
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional

from .path_manager import EXPORTS_DIR, sanitize_filename

logger = logging.getLogger(__name__)

# Check if openpyxl is installed
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl is not installed. Excel export will fall back to CSV.")


def is_excel_available() -> bool:
    """Check if openpyxl library is available."""
    return OPENPYXL_AVAILABLE


def _apply_header_style(ws, start_row: int, end_col: int, fill_color: str = "2563EB", text_color: str = "FFFFFF"):
    """Apply styling to header row."""
    if not OPENPYXL_AVAILABLE:
        return
    header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color=text_color)
    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="medium", color="1E3A8A"),
        bottom=Side(style="medium", color="1E3A8A")
    )
    for col in range(1, end_col + 1):
        cell = ws.cell(row=start_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[start_row].height = 26


def _auto_fit_columns(ws, max_col: int, min_width: int = 12):
    """Auto fit column widths based on maximum string length."""
    if not OPENPYXL_AVAILABLE:
        return
    for col in range(1, max_col + 1):
        col_letter = get_column_letter(col)
        max_len = min_width
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                # String representation length
                line_len = max(len(line) for line in str(val).split("\n"))
                if line_len > max_len:
                    max_len = line_len
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


def export_table_to_excel(
    filepath: str,
    title: str,
    headers: List[str],
    rows: List[List[Any]],
    sheet_name: str = "Report",
    summary_text: str = ""
) -> str:
    """
    Export generic table data to an Excel (.xlsx) file.
    Falls back to CSV if openpyxl is not available or filepath ends in .csv.
    """
    path = Path(filepath)
    if not OPENPYXL_AVAILABLE or path.suffix.lower() == ".csv":
        # Fallback to CSV
        csv_path = path.with_suffix(".csv")
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            if title:
                writer.writerow([title])
                writer.writerow([])
            writer.writerow(headers)
            writer.writerows(rows)
            if summary_text:
                writer.writerow([])
                writer.writerow([summary_text])
        return str(csv_path)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    ws.views.sheetView[0].showGridLines = True

    current_row = 1

    # Title Banner
    if title:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=1, value=title.upper())
        title_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 36
        current_row = 2

        # Subtitle timestamp
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        sub_cell = ws.cell(row=2, column=1, value=f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        sub_cell.font = Font(name="Calibri", size=9, italic=True, color="475569")
        sub_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 20
        current_row = 4

    # Headers
    header_row_idx = current_row
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=header_row_idx, column=col_idx, value=header)
    _apply_header_style(ws, header_row_idx, len(headers))
    current_row += 1

    # Data Rows
    data_font = Font(name="Calibri", size=10)
    alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    grid_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0")
    )

    for r_idx, row_data in enumerate(rows):
        row_num = current_row
        ws.row_dimensions[row_num].height = 22
        fill = alt_fill if r_idx % 2 == 1 else white_fill

        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=c_idx)
            cell.font = data_font
            cell.fill = fill
            cell.border = grid_border

            # Parse value type and formatting
            if isinstance(val, (int, float)):
                cell.value = val
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if isinstance(val, float) or "." in str(val):
                    cell.number_format = "#,##0.00"
                else:
                    cell.number_format = "#,##0"
            elif isinstance(val, str) and (val.startswith("₹ ") or val.startswith("Rs ")):
                try:
                    num_str = val.replace("₹ ", "").replace("Rs ", "").replace(",", "").strip()
                    cell.value = float(num_str)
                    cell.number_format = "₹ #,##0.00"
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                except ValueError:
                    cell.value = val
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.value = str(val) if val is not None else ""
                cell.alignment = Alignment(horizontal="left", vertical="center")

        current_row += 1

    # Summary row
    if summary_text:
        current_row += 1
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
        sum_cell = ws.cell(row=current_row, column=1, value=summary_text)
        sum_cell.font = Font(name="Calibri", size=10, bold=True, color="1E3A8A")
        sum_cell.fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
        sum_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[current_row].height = 26

    _auto_fit_columns(ws, len(headers))

    wb.save(filepath)
    return str(filepath)


def export_all_days_sales_to_excel(
    filepath: Optional[str] = None,
    from_date: str = "",
    to_date: str = ""
) -> str:
    """
    Export EVERY DAY's sales report, bills log, expenses, and summary into a multi-tab Excel Workbook.
    """
    if not filepath:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath = str(EXPORTS_DIR / f"daily_sales_all_days_{stamp}.xlsx")

    path = Path(filepath)
    if not OPENPYXL_AVAILABLE or path.suffix.lower() == ".csv":
        return _export_all_days_csv_fallback(path, from_date, to_date)

    from ..database import dao

    # Fetch all completed bills and expenses
    all_bills = dao.get_all_bills(limit=5000)
    all_expenses = dao.get_all_expenses()
    all_items = dao.get_all_items()

    # Filter bills by date if provided
    filtered_bills = []
    for b in all_bills:
        if b.status != "COMPLETED":
            continue
        b_dt = b.bill_date
        b_day = b_dt.strftime("%Y-%m-%d") if hasattr(b_dt, "strftime") else str(b_dt)[:10]
        if from_date and b_day < from_date:
            continue
        if to_date and b_day > to_date:
            continue
        filtered_bills.append(b)

    # Filter expenses by date if provided
    filtered_expenses = []
    for e in all_expenses:
        e_day = e.expense_date[:10]
        if from_date and e_day < from_date:
            continue
        if to_date and e_day > to_date:
            continue
        filtered_expenses.append(e)

    # Aggregate daily metrics
    from collections import defaultdict
    daily_sales: dict[str, float] = defaultdict(float)
    daily_cash_sales: dict[str, float] = defaultdict(float)
    daily_upi_sales: dict[str, float] = defaultdict(float)
    daily_card_sales: dict[str, float] = defaultdict(float)
    daily_cogs: dict[str, float] = defaultdict(float)
    daily_discount: dict[str, float] = defaultdict(float)
    daily_tax: dict[str, float] = defaultdict(float)
    daily_bill_count: dict[str, int] = defaultdict(int)

    for b in filtered_bills:
        day = b.bill_date.strftime("%Y-%m-%d") if hasattr(b.bill_date, "strftime") else str(b.bill_date)[:10]
        daily_sales[day] += b.total_amount
        daily_bill_count[day] += 1
        daily_discount[day] += getattr(b, "total_discount", getattr(b, "discount", 0.0))
        daily_tax[day] += getattr(b, "tax_amount", 0.0)

        pmode = (b.payment_mode or "CASH").upper()
        if pmode == "CASH":
            daily_cash_sales[day] += b.total_amount
        elif pmode in ("UPI", "ONLINE"):
            daily_upi_sales[day] += b.total_amount
        elif pmode in ("CARD", "CREDIT_CARD"):
            daily_card_sales[day] += b.total_amount
        else:
            daily_cash_sales[day] += b.total_amount

        # COGS calculation
        b_cost = 0.0
        for item in b.bill_items:
            cp = item.purchase_price
            if cp <= 0:
                cp = item.unit_price * 0.60
            b_cost += cp * item.quantity
        daily_cogs[day] += b_cost

    daily_expenses: dict[str, float] = defaultdict(float)
    for e in filtered_expenses:
        e_day = e.expense_date[:10]
        daily_expenses[e_day] += e.amount

    all_days = sorted(set(daily_sales.keys()) | set(daily_expenses.keys()))

    wb = openpyxl.Workbook()

    # ─────────────────────────────────────────────────────────────
    # TAB 1: Daily Sales & Profit Summary (Every Day Row)
    # ─────────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Daily Sales Summary"
    ws1.views.sheetView[0].showGridLines = True

    # Title Banner
    ws1.merge_cells("A1:K1")
    t_cell = ws1["A1"]
    t_cell.value = "BEREEZE FOOTWEAR - DAILY SALES & PROFIT REPORT (EVERY DAY)"
    t_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    t_cell.fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    t_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 36

    ws1.merge_cells("A2:K2")
    sub_cell = ws1["A2"]
    sub_cell.value = f"Date Range: {from_date or 'Beginning'} to {to_date or 'Today'}  |  Exported: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    sub_cell.font = Font(name="Calibri", size=9, italic=True, color="475569")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[2].height = 20

    headers1 = [
        "Date", "Total Bills", "Gross Sales (₹)", "Cash Sales (₹)", "UPI / Online (₹)",
        "Card / Other (₹)", "Discount (₹)", "Cost of Goods (₹)", "Expenses (₹)",
        "Net Profit (₹)", "Profit Margin (%)"
    ]

    for c_idx, h in enumerate(headers1, 1):
        ws1.cell(row=4, column=c_idx, value=h)
    _apply_header_style(ws1, 4, len(headers1), fill_color="2563EB")

    grid_border = Border(
        left=Side(style="thin", color="E2E8F0"), right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"), bottom=Side(style="thin", color="E2E8F0")
    )
    alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    tot_bills = 0
    tot_sales = 0.0
    tot_cash = 0.0
    tot_upi = 0.0
    tot_card = 0.0
    tot_disc = 0.0
    tot_cogs = 0.0
    tot_exp = 0.0

    curr_row = 5
    for r_idx, day_str in enumerate(all_days):
        ws1.row_dimensions[curr_row].height = 22
        fill = alt_fill if r_idx % 2 == 1 else white_fill

        cnt = daily_bill_count[day_str]
        sales = daily_sales[day_str]
        cash = daily_cash_sales[day_str]
        upi = daily_upi_sales[day_str]
        card = daily_card_sales[day_str]
        disc = daily_discount[day_str]
        cogs = daily_cogs[day_str]
        exp = daily_expenses[day_str]
        profit = sales - cogs - exp
        margin = (profit / sales * 100.0) if sales > 0 else 0.0

        tot_bills += cnt
        tot_sales += sales
        tot_cash += cash
        tot_upi += upi
        tot_card += card
        tot_disc += disc
        tot_cogs += cogs
        tot_exp += exp

        row_vals = [day_str, cnt, sales, cash, upi, card, disc, cogs, exp, profit, margin / 100.0]

        for c_idx, val in enumerate(row_vals, 1):
            cell = ws1.cell(row=curr_row, column=c_idx)
            cell.font = Font(name="Calibri", size=10)
            cell.fill = fill
            cell.border = grid_border

            if c_idx == 1:
                cell.value = val
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c_idx == 2:
                cell.value = val
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif c_idx == 11:
                cell.value = val
                cell.number_format = "0.0%"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.value = val
                cell.number_format = "₹ #,##0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")

        curr_row += 1

    # Total Summary Row at bottom
    ws1.row_dimensions[curr_row].height = 26
    tot_profit = tot_sales - tot_cogs - tot_exp
    tot_margin = (tot_profit / tot_sales) if tot_sales > 0 else 0.0

    totals_val = ["TOTAL / SUMMARY", tot_bills, tot_sales, tot_cash, tot_upi, tot_card, tot_disc, tot_cogs, tot_exp, tot_profit, tot_margin]
    tot_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    tot_border = Border(
        top=Side(style="medium", color="1E3A8A"),
        bottom=Side(style="double", color="1E3A8A"),
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1")
    )
    tot_font = Font(name="Calibri", size=10, bold=True, color="1E3A8A")

    for c_idx, val in enumerate(totals_val, 1):
        cell = ws1.cell(row=curr_row, column=c_idx)
        cell.font = tot_font
        cell.fill = tot_fill
        cell.border = tot_border

        if c_idx == 1:
            cell.value = val
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx == 2:
            cell.value = val
            cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="right", vertical="center")
        elif c_idx == 11:
            cell.value = val
            cell.number_format = "0.0%"
            cell.alignment = Alignment(horizontal="right", vertical="center")
        else:
            cell.value = val
            cell.number_format = "₹ #,##0.00"
            cell.alignment = Alignment(horizontal="right", vertical="center")

    _auto_fit_columns(ws1, len(headers1))

    # ─────────────────────────────────────────────────────────────
    # TAB 2: Bills Master Log (Every Bill Detail)
    # ─────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet(title="All Bills Log")
    ws2.views.sheetView[0].showGridLines = True

    ws2.merge_cells("A1:I1")
    t2 = ws2["A1"]
    t2.value = "DETAILED BILLS LOG (ALL TRANSACTIONS)"
    t2.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    t2.fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 32

    headers2 = ["Bill Number", "Date & Time", "Customer Name", "Subtotal (₹)", "Discount (₹)", "Tax (₹)", "Net Total (₹)", "Payment Mode", "Status"]
    for c_idx, h in enumerate(headers2, 1):
        ws2.cell(row=3, column=c_idx, value=h)
    _apply_header_style(ws2, 3, len(headers2), fill_color="0D9488")

    r2 = 4
    for b in filtered_bills:
        ws2.row_dimensions[r2].height = 20
        b_dt_str = b.bill_date.strftime("%Y-%m-%d %H:%M:%S") if hasattr(b.bill_date, "strftime") else str(b.bill_date)
        vals2 = [
            b.bill_number,
            b_dt_str,
            b.customer_name or "Walk-In Customer",
            getattr(b, "subtotal", getattr(b, "sub_total", b.total_amount)),
            getattr(b, "total_discount", getattr(b, "discount", 0.0)),
            getattr(b, "tax_amount", 0.0),
            b.total_amount,
            b.payment_mode or "CASH",
            b.status
        ]

        for c_idx, val in enumerate(vals2, 1):
            cell = ws2.cell(row=r2, column=c_idx)
            cell.font = Font(name="Calibri", size=10)
            cell.border = grid_border

            if c_idx in (4, 5, 6, 7):
                cell.value = float(val or 0)
                cell.number_format = "₹ #,##0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.value = str(val or "")
                cell.alignment = Alignment(horizontal="center" if c_idx in (1, 2, 8, 9) else "left", vertical="center")
        r2 += 1

    _auto_fit_columns(ws2, len(headers2))

    # ─────────────────────────────────────────────────────────────
    # TAB 3: Expenses Log
    # ─────────────────────────────────────────────────────────────
    ws3 = wb.create_sheet(title="Expenses Log")
    ws3.views.sheetView[0].showGridLines = True

    ws3.merge_cells("A1:F1")
    t3 = ws3["A1"]
    t3.value = "STORE EXPENSES RECORD LOG"
    t3.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    t3.fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
    t3.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 32

    headers3 = ["Expense ID", "Date", "Category", "Description", "Amount (₹)", "Payment Mode"]
    for c_idx, h in enumerate(headers3, 1):
        ws3.cell(row=3, column=c_idx, value=h)
    _apply_header_style(ws3, 3, len(headers3), fill_color="DC2626")

    r3 = 4
    for e in filtered_expenses:
        ws3.row_dimensions[r3].height = 20
        vals3 = [e.expense_id, e.expense_date, e.category, e.description, e.amount, e.payment_mode]
        for c_idx, val in enumerate(vals3, 1):
            cell = ws3.cell(row=r3, column=c_idx)
            cell.font = Font(name="Calibri", size=10)
            cell.border = grid_border

            if c_idx == 5:
                cell.value = float(val or 0)
                cell.number_format = "₹ #,##0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif c_idx in (1, 2, 6):
                cell.value = str(val or "")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.value = str(val or "")
                cell.alignment = Alignment(horizontal="left", vertical="center")
        r3 += 1

    _auto_fit_columns(ws3, len(headers3))

    # ─────────────────────────────────────────────────────────────
    # TAB 4: Current Stock & Inventory
    # ─────────────────────────────────────────────────────────────
    ws4 = wb.create_sheet(title="Inventory Stock")
    ws4.views.sheetView[0].showGridLines = True

    ws4.merge_cells("A1:F1")
    t4 = ws4["A1"]
    t4.value = "CURRENT INVENTORY & STOCK VALUATION"
    t4.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    t4.fill = PatternFill(start_color="0284C7", end_color="0284C7", fill_type="solid")
    t4.alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 32

    headers4 = ["Item Code", "Item Name", "Category", "Available Stock", "Selling Price (₹)", "Total Stock Value (₹)"]
    for c_idx, h in enumerate(headers4, 1):
        ws4.cell(row=3, column=c_idx, value=h)
    _apply_header_style(ws4, 3, len(headers4), fill_color="0284C7")

    r4 = 4
    for i in all_items:
        ws4.row_dimensions[r4].height = 20
        stock_val = i.stock_quantity * i.selling_price
        vals4 = [i.item_code, i.item_name, i.category, i.stock_quantity, i.selling_price, stock_val]
        for c_idx, val in enumerate(vals4, 1):
            cell = ws4.cell(row=r4, column=c_idx)
            cell.font = Font(name="Calibri", size=10)
            cell.border = grid_border

            if c_idx == 4:
                cell.value = int(val)
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif c_idx in (5, 6):
                cell.value = float(val)
                cell.number_format = "₹ #,##0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.value = str(val or "")
                cell.alignment = Alignment(horizontal="left", vertical="center")
        r4 += 1

    _auto_fit_columns(ws4, len(headers4))

    wb.save(filepath)
    return str(filepath)


def export_day_closing_to_excel(filepath: str, closing_data: dict) -> str:
    """Export Day Closing / Shop End Report into a beautifully structured Excel file."""
    path = Path(filepath)
    if not OPENPYXL_AVAILABLE or path.suffix.lower() == ".csv":
        # CSV Fallback
        csv_path = path.with_suffix(".csv")
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["BEREEZE FOOTWEAR - SHOP CLOSING & DAY END REPORT"])
            writer.writerow(["Date", closing_data.get("closing_date", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))])
            writer.writerow(["Closed By", closing_data.get("user", "Admin")])
            writer.writerow([])
            writer.writerow(["METRIC", f"YESTERDAY ({closing_data.get('yesterday_str','')})", f"TODAY ({closing_data.get('today_str','')})"])
            y_data = closing_data.get("yesterday_data", {})
            t_data = closing_data.get("today_data", {})
            writer.writerow(["Gross Cash Sales", f"{y_data.get('cash_sales',0):.2f}", f"{t_data.get('cash_sales',0):.2f}"])
            writer.writerow(["UPI / Online Sales", f"{y_data.get('upi_sales',0):.2f}", f"{t_data.get('upi_sales',0):.2f}"])
            writer.writerow(["Total Expenses", f"{y_data.get('total_expenses',0):.2f}", f"{t_data.get('total_expenses',0):.2f}"])
            writer.writerow(["Net Cash in Counter", f"{y_data.get('net_cash_in_counter',0):.2f}", f"{t_data.get('net_cash_in_counter',0):.2f}"])
            writer.writerow([])
            writer.writerow(["CASH TALLY VERIFICATION"])
            writer.writerow(["Physical Cash Counted", closing_data.get("cash_counted", "0.00")])
            writer.writerow(["Variance", closing_data.get("variance_text", "Matched")])
            writer.writerow(["Closing Remarks", closing_data.get("remarks", "")])
        return str(csv_path)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Day Closing Report"
    ws.views.sheetView[0].showGridLines = True

    # Title
    ws.merge_cells("A1:C1")
    t = ws["A1"]
    t.value = "BEREEZE FOOTWEAR - SHOP CLOSING & DAY END REPORT"
    t.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    t.fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # Meta
    ws.cell(row=2, column=1, value="Closing Date & Time:").font = Font(name="Calibri", size=10, bold=True)
    ws.cell(row=2, column=2, value=closing_data.get("closing_date", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))).font = Font(name="Calibri", size=10)

    ws.cell(row=3, column=1, value="Closed By:").font = Font(name="Calibri", size=10, bold=True)
    ws.cell(row=3, column=2, value=closing_data.get("user", "Admin")).font = Font(name="Calibri", size=10)

    # Section 1: Metrics
    ws.cell(row=5, column=1, value="METRIC").font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws.cell(row=5, column=2, value=f"YESTERDAY ({closing_data.get('yesterday_str','')})").font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws.cell(row=5, column=3, value=f"TODAY ({closing_data.get('today_str','')})").font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    _apply_header_style(ws, 5, 3, fill_color="2563EB")

    y_data = closing_data.get("yesterday_data", {})
    t_data = closing_data.get("today_data", {})

    metrics_rows = [
        ("Gross Cash Sales (₹)", y_data.get("cash_sales", 0.0), t_data.get("cash_sales", 0.0)),
        ("UPI / Online Sales (₹)", y_data.get("upi_sales", 0.0), t_data.get("upi_sales", 0.0)),
        ("Total Expenses (₹)", y_data.get("total_expenses", 0.0), t_data.get("total_expenses", 0.0)),
        ("Net Cash in Counter (₹)", y_data.get("net_cash_in_counter", 0.0), t_data.get("net_cash_in_counter", 0.0)),
    ]

    grid_border = Border(
        left=Side(style="thin", color="E2E8F0"), right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"), bottom=Side(style="thin", color="E2E8F0")
    )

    r = 6
    for label, y_val, t_val in metrics_rows:
        ws.row_dimensions[r].height = 22
        c1 = ws.cell(row=r, column=1, value=label)
        c2 = ws.cell(row=r, column=2, value=float(y_val))
        c3 = ws.cell(row=r, column=3, value=float(t_val))

        c1.font = Font(name="Calibri", size=10, bold=True)
        c2.font = Font(name="Calibri", size=10)
        c3.font = Font(name="Calibri", size=10, bold=True)

        c1.border = grid_border
        c2.border = grid_border
        c3.border = grid_border

        c2.number_format = "₹ #,##0.00"
        c3.number_format = "₹ #,##0.00"
        c2.alignment = Alignment(horizontal="right", vertical="center")
        c3.alignment = Alignment(horizontal="right", vertical="center")
        r += 1

    # Section 2: Cash Tally
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    ws.cell(row=r, column=1, value="CASH TALLY VERIFICATION").font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    _apply_header_style(ws, r, 3, fill_color="0D9488")
    r += 1

    tally_rows = [
        ("Physical Cash Counted:", closing_data.get("cash_counted", "0.00")),
        ("Variance Status:", closing_data.get("variance_text", "Matched")),
        ("Closing Remarks:", closing_data.get("remarks", "")),
    ]

    for lbl, val in tally_rows:
        ws.row_dimensions[r].height = 22
        ws.cell(row=r, column=1, value=lbl).font = Font(name="Calibri", size=10, bold=True)
        ws.cell(row=r, column=1).border = grid_border
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        vc = ws.cell(row=r, column=2, value=str(val))
        vc.font = Font(name="Calibri", size=10)
        vc.border = grid_border
        ws.cell(row=r, column=3).border = grid_border
        r += 1

    _auto_fit_columns(ws, 3)
    wb.save(filepath)
    return str(filepath)


def _export_all_days_csv_fallback(filepath: Path, from_date: str, to_date: str) -> str:
    """Fallback CSV exporter for all days sales data when openpyxl is unavailable."""
    from ..database import dao

    csv_path = filepath.with_suffix(".csv")
    all_bills = dao.get_all_bills(limit=5000)
    all_expenses = dao.get_all_expenses()

    from collections import defaultdict
    daily_sales = defaultdict(float)
    daily_count = defaultdict(int)
    daily_exp = defaultdict(float)

    for b in all_bills:
        if b.status != "COMPLETED":
            continue
        day = b.bill_date.strftime("%Y-%m-%d") if hasattr(b.bill_date, "strftime") else str(b.bill_date)[:10]
        if from_date and day < from_date:
            continue
        if to_date and day > to_date:
            continue
        daily_sales[day] += b.total_amount
        daily_count[day] += 1

    for e in all_expenses:
        day = e.expense_date[:10]
        if from_date and day < from_date:
            continue
        if to_date and day > to_date:
            continue
        daily_exp[day] += e.amount

    all_days = sorted(set(daily_sales.keys()) | set(daily_exp.keys()))

    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["BEREEZE FOOTWEAR - DAILY SALES REPORT (EVERY DAY)"])
        writer.writerow(["Exported", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        writer.writerow([])
        writer.writerow(["Date", "Total Bills", "Gross Sales", "Expenses", "Net Revenue"])
        for day in all_days:
            s = daily_sales[day]
            ex = daily_exp[day]
            writer.writerow([day, daily_count[day], f"{s:.2f}", f"{ex:.2f}", f"{s - ex:.2f}"])

    return str(csv_path)
